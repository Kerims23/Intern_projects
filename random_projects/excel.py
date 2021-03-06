import openpyxl as xl
from openpyxl.chart import bar_chart, Reference

#creating a def statment to load in file
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['sheet1']

#adding a for loop that goes through the file
    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row,3)
        corrected_price=float(cell.value.replace('$',''))*0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    values = reference(sheet, min_row=2,max_row=sheet.max_row, min_col=4)=bar_chart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')
    wb.save(filename)

